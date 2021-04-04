import os

import openpyxl
import datetime
import random

import os
import subprocess

from win32com import client


class ExcelData:
    def __init__(self, dataTime: datetime.date,
                 actNumber: int,
                 mol: str,
                 worker: str,
                 object: str,
                 positions: dict):
        self.dataTime = dataTime
        self.actNumber = actNumber
        self.mol = mol
        self.worker = worker
        self.object = object
        self.positions = positions

    @staticmethod
    def dict_to_class(dict_):
        return ExcelData(
            dataTime=dict_["dataTime"],
            actNumber=dict_["actNumber"],
            mol=dict_["mol"],
            worker=dict_["worker"],
            object=dict_["object"],
            positions=dict_["positions"]
        )


path = os.getcwd() + "\\Positions\\prin-template.xlsx"

testData = {
    "dataTime": datetime.date.today(),
    "actNumber": random.randint(0, 100000),
    "mol": "Тест Ответтственный",
    "worker": "Тест приемщик",
    "object": "Тест объект",
    "positions": [
        {
            "name": "test1",
            "code": random.randint(0, 1000000),
            "ediz": "шт",
            "quantity": random.randint(1, 100)
        },
        {
            "name": "test2",
            "code": random.randint(0, 1000000),
            "ediz": "шт",
            "quantity": random.randint(1, 100)
        }],
}

data = ExcelData.dict_to_class(testData)


def write_to_excel(data: ExcelData):
    print(path)
    wb = openpyxl.load_workbook(path)
    xl = wb.active

    datime = xl.cell(3, 8)
    ID = xl.cell(3, 5)
    mol = xl.cell(6, 2)
    prin = xl.cell(9, 2)
    obj = xl.cell(12, 2)

    datime.value = data.dataTime
    ID.value = data.actNumber
    mol.value = data.mol
    prin.value = data.worker
    obj.value = data.object

    first = 20
    # first pos

    for i, k in enumerate(data.positions):
        posID = xl.cell(first + i, 2)
        posName = xl.cell(first + i, 4)
        posCode = xl.cell(first + i, 5)
        posEdiz = xl.cell(first + i, 7)
        posQuant = xl.cell(first + i, 12)

        posID.value = i + 1
        posName.value = k["name"]
        posCode.value = k["code"]
        posEdiz.value = k["ediz"]
        posQuant.value = k["quantity"]

    if os.path.isfile("1.xlsx"):
        os.remove("1.xlsx")
    wb.save("1.xlsx")


def excel_to_pdf():
    try:
        xlApp = client.Dispatch("Excel.Application")
    except Exception as ex:
        return ("Приложение Excel недоступно\n", ex)
    books = xlApp.Workbooks.Open(os.getcwd() + '\\1.xlsx')
    try:
        ws = books.Worksheets[0]
        ws.Visible = 1
        ws.ExportAsFixedFormat(0, os.getcwd() + '\\1')
    except Exception as ex:
        books.Close()
        return ("Експорт в ПДФ неудачный", ex)
    books.Close()


def print_pdf():
    args = '"C:\\\\Program Files\\\\gs\\\\gs9.54.0\\\\bin\\\\gswin64c" ' \
           '-sDEVICE=mswinpr2 ' \
           '-dBATCH ' \
           '-dNOPAUSE ' \
           '-dFitPage ' \
           '-sOutputFile="%printer%myPrinterName" '
    ghostscript = args + os.path.join(os.getcwd(), '1.pdf').replace('\\', '\\\\')
    subprocess.call(ghostscript, shell=False)


if __name__ == "__main__":
    path = "prin-template.xlsx"
    write_to_excel(data)
    excel_to_pdf()
